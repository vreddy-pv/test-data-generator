import openpyxl

class OutputWriter:
    def __init__(self, generated_data):
        self.generated_data = generated_data

    def write_sql_inserts(self, output_file):
        with open(output_file, 'w') as f:
            for table_name, rows in self.generated_data.items():
                if not rows:
                    continue
                
                columns = rows[0].keys()
                f.write(f"-- Data for table: {table_name}\n")
                for row in rows:
                    values = []
                    for col in columns:
                        value = row[col]
                        if isinstance(value, str):
                            # Simple escape for single quotes
                            value = value.replace("'", "''")
                            values.append(f"'{value}'")
                        elif value is None:
                            values.append("NULL")
                        else:
                            values.append(str(value))
                    
                    f.write(f"INSERT INTO {table_name} ({', '.join(columns)}) VALUES ({', '.join(values)});\n")
                f.write("\n")

    def write_excel(self, output_file):
        wb = openpyxl.Workbook()
        # Remove the default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        for table_name, rows in self.generated_data.items():
            ws = wb.create_sheet(title=table_name)
            
            if not rows:
                continue

            headers = list(rows[0].keys())
            ws.append(headers)

            for row in rows:
                ws.append([row[h] for h in headers])
        
        wb.save(output_file)
